"""
Wily Facturador - App Web
Factura C para cuenta personal (Monotributista)
"""

from flask import Flask, render_template, request, redirect, url_for, session, flash
from functools import wraps
import json
import os
import csv
import io
from datetime import datetime, timezone, timedelta
import psycopg2
from psycopg2.extras import RealDictCursor

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "wily-secret-2024")

DATABASE_URL = os.environ.get("DATABASE_URL")

USUARIOS = {
    "facturador": {
        "password": os.environ.get("PASS_FACTURADOR", "wily123"),
        "nombre": "Facturador",
    },
    "admin": {
        "password": os.environ.get("PASS_ADMIN", "admin2024"),
        "nombre": "Admin",
    },
}

PUNTO_VENTA = int(os.environ.get("PUNTO_VENTA", "1"))
TIPO_COMPROBANTE = 11  # Factura C

def hora_argentina():
    """Devuelve el datetime actual en horario de Argentina (UTC-3)."""
    return datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=-3)))

def get_conn():
    """Conexion a Postgres (Supabase). Cada fila se devuelve como diccionario."""
    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)

def init_db():
    if not DATABASE_URL:
        print("[AVISO] DATABASE_URL no esta configurada. La app no podra guardar datos.")
        return
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS comprobantes (
                id           SERIAL PRIMARY KEY,
                numero       INTEGER,
                punto_venta  INTEGER,
                fecha        TEXT,
                destinatario TEXT,
                doc_tipo     INTEGER,
                doc_nro      TEXT,
                monto        DOUBLE PRECISION,
                cae          TEXT,
                cae_vto      TEXT,
                estado       TEXT,
                error        TEXT,
                usuario      TEXT,
                created_at   TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS logs (
                id         SERIAL PRIMARY KEY,
                nivel      TEXT,
                mensaje    TEXT,
                usuario    TEXT,
                created_at TEXT
            )
        """)
        conn.commit()
        cur.close()
        conn.close()
        print("[INFO] Base de datos Postgres lista")
    except Exception as e:
        print("[ERROR] No se pudo inicializar la base de datos:", str(e))

init_db()

def log_evento(nivel, mensaje, usuario=None):
    """Guarda un evento en la tabla de logs. nivel: INFO / ERROR / WARN"""
    hora_str = hora_argentina().strftime("%Y-%m-%d %H:%M:%S")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO logs (nivel, mensaje, usuario, created_at) VALUES (%s,%s,%s,%s)",
            (nivel, mensaje, usuario or session.get("usuario", "sistema"), hora_str)
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print("[ERROR] No se pudo guardar el log:", str(e))
    print("[{}] {}".format(nivel, mensaje))

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "usuario" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        usuario = request.form.get("usuario", "").lower()
        password = request.form.get("password", "")
        if usuario in USUARIOS and USUARIOS[usuario]["password"] == password:
            session["usuario"] = usuario
            session["nombre"] = USUARIOS[usuario]["nombre"]
            log_evento("INFO", "Inicio de sesion exitoso", usuario)
            return redirect(url_for("dashboard"))
        error = "Usuario o contrasena incorrectos"
        log_evento("WARN", "Intento de login fallido con usuario '{}'".format(usuario), usuario)
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/")
@login_required
def dashboard():
    mes_actual = hora_argentina().strftime("%Y-%m")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM comprobantes ORDER BY created_at DESC LIMIT 100")
    rows = cur.fetchall()
    cur.execute("""
        SELECT COALESCE(SUM(monto),0) AS total FROM comprobantes
        WHERE estado='APROBADO' AND LEFT(fecha, 7) = %s
    """, (mes_actual,))
    total_mes = cur.fetchone()["total"]
    cur.execute("""
        SELECT COUNT(*) AS cant FROM comprobantes
        WHERE estado='APROBADO' AND LEFT(fecha, 7) = %s
    """, (mes_actual,))
    cant_mes = cur.fetchone()["cant"]
    cur.close()
    conn.close()

    return render_template("dashboard.html",
        comprobantes=rows,
        total_mes=total_mes,
        cant_mes=cant_mes,
        usuario=session["usuario"],
    )

@app.route("/emitir", methods=["GET", "POST"])
@login_required
def emitir():
    if request.method == "POST":
        data = {
            "nombre":   request.form.get("nombre"),
            "doc_tipo": int(request.form.get("doc_tipo", 99)),
            "doc_nro":  request.form.get("doc_nro", ""),
            "monto":    float(request.form.get("monto", 0)),
        }
        resultado = emitir_factura_arca(data)
        guardar_comprobante(resultado)

        if resultado["estado"] == "APROBADO":
            flash("Factura C #{} emitida. CAE: {}".format(
                resultado["numero"], resultado["cae"]), "success")
            log_evento("INFO", "Factura C #{} emitida a '{}' por ${:,.2f}. CAE: {}".format(
                resultado["numero"], resultado["destinatario"], resultado["monto"], resultado["cae"]))
        else:
            flash("Error: {}".format(resultado["error"]), "error")
            log_evento("ERROR", "Factura rechazada para '{}' por ${:,.2f}. Motivo: {}".format(
                data["nombre"], data["monto"], resultado["error"]))

        return redirect(url_for("dashboard"))

    return render_template("emitir.html", usuario=session["usuario"])

def _parsear_archivo_masivo(archivo):
    """Lee un Excel o CSV y devuelve una lista de filas normalizadas
    con las claves: nombre, doc_tipo, doc_nro, monto"""
    nombre_archivo = archivo.filename.lower()
    filas = []

    if nombre_archivo.endswith(".xlsx") or nombre_archivo.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active

        header_row = None
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            valores = [str(v).strip().upper() if v else "" for v in row]
            if any(v in ("RAZONSOCIAL", "NOMBRE", "TIPODOCUMENTO", "DOC_TIPO") for v in valores):
                header_row = i
                headers = [str(v).strip().lower() if v else "" for v in row]
                break

        if not header_row:
            raise ValueError("No se encontraron los encabezados en el Excel.")

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(row):
                continue
            fila = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
            filas.append(fila)
    else:
        content = archivo.read().decode("utf-8")
        reader = csv.DictReader(io.StringIO(content))
        filas = [{k.lower().strip(): v for k, v in row.items()} for row in reader]

    normalizadas = []
    for fila in filas:
        nombre = (fila.get("razonsocial") or fila.get("nombre") or "").strip()
        monto = float(fila.get("monto") or 0)

        tipo_doc_raw = (fila.get("tipodocumento") or fila.get("doc_tipo") or "CF").strip().upper()
        if tipo_doc_raw in ("CUIT", "80"):
            doc_tipo = 80
        elif tipo_doc_raw in ("DNI", "96"):
            doc_tipo = 96
        else:
            doc_tipo = 99

        doc_nro = (fila.get("nrodocumento") or fila.get("doc_nro") or "").strip()
        doc_nro = doc_nro.replace("-", "").replace(".", "")

        normalizadas.append({
            "nombre":   nombre,
            "doc_tipo": doc_tipo,
            "doc_nro":  doc_nro,
            "monto":    monto,
        })
    return normalizadas

@app.route("/masivo", methods=["GET"])
@login_required
def masivo():
    return render_template("masivo.html", usuario=session["usuario"])

@app.route("/masivo/preview", methods=["POST"])
@login_required
def masivo_preview():
    archivo = request.files.get("archivo")
    if not archivo:
        return {"ok": False, "error": "No se selecciono ningun archivo"}, 400
    try:
        filas = _parsear_archivo_masivo(archivo)
        if not filas:
            return {"ok": False, "error": "El archivo no tiene filas de datos"}, 400
        log_evento("INFO", "Carga masiva: archivo '{}' cargado con {} filas".format(
            archivo.filename, len(filas)))
        return {"ok": True, "filas": filas, "total": len(filas)}
    except Exception as e:
        log_evento("ERROR", "Error al leer archivo de carga masiva: {}".format(str(e)))
        return {"ok": False, "error": str(e)}, 400

@app.route("/masivo/emitir_uno", methods=["POST"])
@login_required
def masivo_emitir_uno():
    data = request.get_json(force=True)
    fila_data = {
        "nombre":   data.get("nombre", ""),
        "doc_tipo": int(data.get("doc_tipo", 99)),
        "doc_nro":  data.get("doc_nro", ""),
        "monto":    float(data.get("monto", 0)),
    }
    resultado = emitir_factura_arca(fila_data)
    guardar_comprobante(resultado)

    if resultado["estado"] == "APROBADO":
        log_evento("INFO", "[Masivo] Factura C #{} emitida a '{}' por ${:,.2f}. CAE: {}".format(
            resultado["numero"], resultado["destinatario"], resultado["monto"], resultado["cae"]))
    else:
        log_evento("ERROR", "[Masivo] Factura rechazada para '{}' por ${:,.2f}. Motivo: {}".format(
            fila_data["nombre"], fila_data["monto"], resultado["error"]))

    return {"ok": True, "resultado": resultado}

@app.route("/logs")
@login_required
def logs():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM logs ORDER BY created_at DESC LIMIT 300")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("logs.html", logs=rows, usuario=session["usuario"])

def emitir_factura_arca(data):
    import random, string

    MODO_DEMO = not os.environ.get("CUIT")
    hoy = hora_argentina().strftime("%Y-%m-%d")

    if MODO_DEMO:
        numero = random.randint(1000, 9999)
        cae = "".join(random.choices(string.digits, k=14))
        return {
            "numero":        numero,
            "punto_venta":   PUNTO_VENTA,
            "fecha":         hoy,
            "destinatario":  data["nombre"],
            "doc_tipo":      data["doc_tipo"],
            "doc_nro":       data["doc_nro"],
            "monto":         data["monto"],
            "cae":           cae,
            "cae_vto":       "20251231",
            "estado":        "APROBADO",
            "error":         None,
            "usuario":       session["usuario"],
        }

    try:
        import urllib3
        import requests
        from requests.adapters import HTTPAdapter
        from urllib3.util.ssl_ import create_urllib3_context
        import ssl

        class DHAdapter(HTTPAdapter):
            def init_poolmanager(self, *args, **kwargs):
                ctx = create_urllib3_context()
                ctx.set_ciphers("DEFAULT:@SECLEVEL=1")
                ctx.options |= ssl.OP_LEGACY_SERVER_CONNECT
                kwargs["ssl_context"] = ctx
                return super().init_poolmanager(*args, **kwargs)

        session_requests = requests.Session()
        session_requests.mount("https://", DHAdapter())

        from zeep import Client
        from zeep.transports import Transport
        import tempfile
        import xml.etree.ElementTree as ET
        from datetime import timedelta

        CUIT       = os.environ["CUIT"]
        CERT_PATH  = os.environ["CERT_PATH"]
        KEY_PATH   = os.environ["KEY_PATH"]
        PRODUCCION = os.environ.get("PRODUCCION", "false").lower() == "true"
        CONCEPTO   = int(os.environ.get("CONCEPTO", "2"))

        WSAA_URL = (
            "https://wsaa.afip.gov.ar/ws/services/LoginCms?wsdl"
            if PRODUCCION else
            "https://wsaahomo.afip.gov.ar/ws/services/LoginCms?wsdl"
        )
        WSFE_URL = (
            "https://servicios1.afip.gov.ar/wsfev1/service.asmx?WSDL"
            if PRODUCCION else
            "https://wswhomo.afip.gov.ar/wsfev1/service.asmx?WSDL"
        )

        cache_file = "/tmp/ticket_{}.json".format(CUIT)
        ticket = None
        if os.path.exists(cache_file):
            cached = json.loads(open(cache_file).read())
            from datetime import timezone
            now_utc = datetime.now(timezone.utc)
            expiry = datetime.fromisoformat(cached["expiry"])
            if expiry.tzinfo is None:
                expiry = expiry.replace(tzinfo=timezone.utc)
            if now_utc < expiry - timedelta(minutes=10):
                ticket = cached

        if not ticket:
            from datetime import timezone
            now = datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=-3)))
            exp = now + timedelta(hours=12)
            xml_tra = """<?xml version="1.0" encoding="UTF-8"?>
<loginTicketRequest version="1.0">
  <header>
    <uniqueId>{}</uniqueId>
    <generationTime>{}</generationTime>
    <expirationTime>{}</expirationTime>
  </header>
  <service>wsfe</service>
</loginTicketRequest>""".format(
                int(now.timestamp()),
                now.strftime('%Y-%m-%dT%H:%M:%S'),
                exp.strftime('%Y-%m-%dT%H:%M:%S')
            ).encode()

            from cryptography.hazmat.primitives.serialization.pkcs7 import PKCS7SignatureBuilder
            from cryptography.hazmat.primitives.serialization import Encoding
            from cryptography.hazmat.primitives import hashes
            from cryptography.hazmat.backends import default_backend
            from cryptography.hazmat.primitives.serialization import load_pem_private_key
            from cryptography.x509 import load_pem_x509_certificate
            import base64

            with open(CERT_PATH, "rb") as fc:
                cert = load_pem_x509_certificate(fc.read(), default_backend())
            with open(KEY_PATH, "rb") as fk:
                key = load_pem_private_key(fk.read(), password=None, backend=default_backend())

            cms_der = (
                PKCS7SignatureBuilder()
                .set_data(xml_tra)
                .add_signer(cert, key, hashes.SHA256())
                .sign(Encoding.DER, options=[])
            )
            cms_b64 = base64.b64encode(cms_der).decode()

            wsaa = Client(WSAA_URL, transport=Transport(session=session_requests))
            resp_xml = wsaa.service.loginCms(in0=cms_b64)
            root = ET.fromstring(resp_xml)

            ticket = {
                "token":  root.find(".//token").text,
                "sign":   root.find(".//sign").text,
                "expiry": exp.isoformat(),
            }
            open(cache_file, "w").write(json.dumps(ticket))

        auth = {"Token": ticket["token"], "Sign": ticket["sign"], "Cuit": int(CUIT)}

        wsfe = Client(WSFE_URL, transport=Transport(session=session_requests))
        ultimo = wsfe.service.FECompUltimoAutorizado(
            Auth=auth, PtoVta=PUNTO_VENTA, CbteTipo=TIPO_COMPROBANTE
        ).CbteNro
        numero = ultimo + 1

        hoy_arca = datetime.now().strftime("%Y%m%d")
        monto = round(data["monto"], 2)

        detalle = {
            "Concepto":     CONCEPTO,
            "DocTipo":      data["doc_tipo"],
            "DocNro":       int(data["doc_nro"]) if data["doc_nro"] else 0,
            "CbteDesde":    numero,
            "CbteHasta":    numero,
            "CbteFch":      hoy_arca,
            "ImpTotal":     monto,
            "ImpTotConc":   0,
            "ImpNeto":      monto,
            "ImpOpEx":      0,
            "ImpIVA":       0,
            "ImpTrib":      0,
            "FchServDesde": hoy_arca if CONCEPTO in (2, 3) else None,
            "FchServHasta": hoy_arca if CONCEPTO in (2, 3) else None,
            "FchVtoPago":   hoy_arca if CONCEPTO in (2, 3) else None,
            "MonId":        "PES",
            "MonCotiz":     1,
        }

        resp = wsfe.service.FECAESolicitar(Auth=auth, FeCAEReq={
            "FeCabReq": {"CantReg": 1, "PtoVta": PUNTO_VENTA, "CbteTipo": TIPO_COMPROBANTE},
            "FeDetReq": {"FECAEDetRequest": [detalle]},
        })

        det = resp.FeDetResp.FECAEDetResponse[0]
        if det.Resultado == "A":
            return {
                "numero": numero, "punto_venta": PUNTO_VENTA,
                "fecha": hoy, "destinatario": data["nombre"],
                "doc_tipo": data["doc_tipo"], "doc_nro": data["doc_nro"],
                "monto": monto, "cae": det.CAE, "cae_vto": det.CAEFchVto,
                "estado": "APROBADO", "error": None,
                "usuario": session["usuario"],
            }
        else:
            errores = []
            if resp.Errors:
                for e in resp.Errors.Err:
                    errores.append("[E{}] {}".format(e.Code, e.Msg))
            if det.Observaciones:
                for o in det.Observaciones.Obs:
                    errores.append("[O{}] {}".format(o.Code, o.Msg))
            raise Exception("; ".join(errores) if errores else "ARCA rechazo sin detalle. Resultado: " + str(det.Resultado))

    except Exception as e:
        import traceback
        error_detalle = traceback.format_exc()
        error_msg = str(e) if str(e) else "Error desconocido: " + type(e).__name__ + " - " + error_detalle[-500:]
        log_evento("ERROR", "Excepcion tecnica al emitir factura: {}".format(error_msg))
        return {
            "numero": None, "punto_venta": PUNTO_VENTA,
            "fecha": hoy, "destinatario": data["nombre"],
            "doc_tipo": data.get("doc_tipo"), "doc_nro": data.get("doc_nro"),
            "monto": data["monto"], "cae": None, "cae_vto": None,
            "estado": "RECHAZADO", "error": error_msg,
            "usuario": session["usuario"],
        }

def guardar_comprobante(r):
    hora_str = hora_argentina().strftime("%Y-%m-%d %H:%M:%S")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO comprobantes
            (numero, punto_venta, fecha, destinatario, doc_tipo, doc_nro,
             monto, cae, cae_vto, estado, error, usuario, created_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            r.get("numero"), r.get("punto_venta"), r.get("fecha"),
            r.get("destinatario"), r.get("doc_tipo"), r.get("doc_nro"),
            r.get("monto"), r.get("cae"), r.get("cae_vto"),
            r.get("estado"), r.get("error"), r.get("usuario"), hora_str,
        ))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        log_evento("ERROR", "No se pudo guardar el comprobante en la base de datos: {}".format(str(e)))

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
